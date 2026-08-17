#!/usr/bin/env python3
"""
xlsx_generator.py — Generador de Modelo Financiero Excel para el Dimensionador Estratégico.

Genera un archivo .xlsx estructurado con diseño corporativo Iris StartUp Lab (Design_2.md)
y gráficos nativos de Excel (Score /25, Trayectoria de Ingresos a 5 años y TAM/SAM/SOM).

Uso:
    python scripts/xlsx_generator.py [--data report_data.json] [--output modelo_financiero.xlsx]
"""

import sys
import json
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

# Paleta de Colores Iris StartUp Lab
HEX_PURPLE_DARK = "241B33"
HEX_PURPLE_LIGHT = "F7F3FC"
HEX_PURPLE_LINE = "E4DCEF"
HEX_GOLD = "D4A73E"
HEX_GOLD_LIGHT = "FEF9C3"
HEX_GREEN = "15803D"
HEX_GREEN_LIGHT = "DCFCE7"
HEX_RED = "B84A3D"
HEX_RED_LIGHT = "FEE2E2"

# Estilos reutilizables
font_title = Font(name="Calibri", size=14, bold=True, color="241B33")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_body = Font(name="Calibri", size=10, bold=False, color="241B33")
font_bold = Font(name="Calibri", size=10, bold=True, color="241B33")

fill_header = PatternFill(start_color=HEX_PURPLE_DARK, end_color=HEX_PURPLE_DARK, fill_type="solid")
fill_zebra = PatternFill(start_color=HEX_PURPLE_LIGHT, end_color=HEX_PURPLE_LIGHT, fill_type="solid")

fill_prototipar = PatternFill(start_color=HEX_GREEN_LIGHT, end_color=HEX_GREEN_LIGHT, fill_type="solid")
font_prototipar = Font(name="Calibri", size=10, bold=True, color=HEX_GREEN)

fill_validar = PatternFill(start_color=HEX_GOLD_LIGHT, end_color=HEX_GOLD_LIGHT, fill_type="solid")
font_validar = Font(name="Calibri", size=10, bold=True, color="92400E")

fill_descartar = PatternFill(start_color=HEX_RED_LIGHT, end_color=HEX_RED_LIGHT, fill_type="solid")
font_descartar = Font(name="Calibri", size=10, bold=True, color=HEX_RED)

thin_border = Border(
    left=Side(style='thin', color=HEX_PURPLE_LINE),
    right=Side(style='thin', color=HEX_PURPLE_LINE),
    top=Side(style='thin', color=HEX_PURPLE_LINE),
    bottom=Side(style='thin', color=HEX_PURPLE_LINE)
)

def style_range(ws, min_col, min_row, max_col, max_row, font=font_body, fill=None, border=thin_border, alignment=None):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if font: cell.font = font
            if fill: cell.fill = fill
            if border: cell.border = border
            if alignment: cell.alignment = alignment

def auto_fit_columns(ws):
    for col in ws.columns:
        first_cell = col[0]
        col_letter = get_column_letter(first_cell.column)
        max_len = 0
        for cell in col:
            if cell.value is not None:
                val_str = str(cell.value)
                if '\n' in val_str:
                    val_str = max(val_str.split('\n'), key=len)
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

def generate_excel_model(data, output_path="modelo_financiero_dimensionamiento.xlsx"):
    wb = openpyxl.Workbook()

    ideas = data.get("ideas", [])

    # -------------------------------------------------------------
    # PESTAÑA 1: 01_Resumen_Priorizacion
    # -------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "01_Resumen_Priorizacion"

    ws1.merge_cells("A1:J1")
    ws1["A1"] = "DIMENSIONADOR ESTRATÉGICO — MATRIZ DE PRIORIZACIÓN DE IDEAS"
    ws1["A1"].font = font_title
    ws1["A1"].alignment = Alignment(vertical="center")
    ws1.row_dimensions[1].height = 30

    headers_ws1 = ["Rank", "Idea de Innovación", "Modelo / Tipo", "Score /25", "SOM 1 Año", "SOM 3 Años", "SOM 5 Años", "CLV:CAC Ajustado", "Buyer Personas", "Veredicto Final"]
    ws1.append(headers_ws1)
    ws1.row_dimensions[2].height = 24

    for col_num, h_text in enumerate(headers_ws1, 1):
        cell = ws1.cell(row=2, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for idx, idea in enumerate(ideas, 1):
        row_num = idx + 2
        verdict = idea.get("verdict", "")
        
        ws1.append([
            f"#{idx}",
            idea.get("name", ""),
            idea.get("model", ""),
            idea.get("score", 0),
            idea.get("som1y", idea.get("som3y", "")),
            idea.get("som3y", ""),
            idea.get("som5y", idea.get("som3y", "")),
            idea.get("clvCacAdjusted", ""),
            ", ".join(idea.get("buyerPersonas", [])),
            verdict
        ])
        
        ws1.row_dimensions[row_num].height = 20
        fill_row = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
        style_range(ws1, min_col=1, min_row=row_num, max_col=10, max_row=row_num, font=font_body, fill=fill_row)
        
        # Formato de veredicto
        v_cell = ws1.cell(row=row_num, column=10)
        v_cell.alignment = Alignment(horizontal="center", vertical="center")
        verdict_norm = verdict.upper().replace('\u00c1', 'A').replace('\u00e1', 'a')  # normaliza Á/á
        if verdict_norm == "PROTOTIPAR":
            v_cell.fill, v_cell.font = fill_prototipar, font_prototipar
        elif "VALIDAR" in verdict_norm:
            v_cell.fill, v_cell.font = fill_validar, font_validar
        else:
            v_cell.fill, v_cell.font = fill_descartar, font_descartar

        ws1.cell(row=row_num, column=4).alignment = Alignment(horizontal="center")
        ws1.cell(row=row_num, column=8).alignment = Alignment(horizontal="center")

    # Gráfico Nativo 1: Score por Idea
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Score de Atractivo (/25) por Idea"
    chart1.y_axis.title = "Score /25"
    chart1.x_axis.title = "Ideas"
    chart1.height = 12
    chart1.width = 18

    data1 = Reference(ws1, min_col=4, min_row=2, max_row=len(ideas)+2)
    cats1 = Reference(ws1, min_col=2, min_row=3, max_row=len(ideas)+2)
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    ws1.add_chart(chart1, "L2")

    auto_fit_columns(ws1)

    # -------------------------------------------------------------
    # PESTAÑA 2: 02_Proyeccion_Ingresos
    # -------------------------------------------------------------
    ws2 = wb.create_sheet(title="02_Proyeccion_Ingresos")
    ws2.merge_cells("A1:I1")
    ws2["A1"] = "PROYECCIÓN DE INGRESOS A 5 AÑOS (USD)"
    ws2["A1"].font = font_title
    ws2["A1"].alignment = Alignment(vertical="center")
    ws2.row_dimensions[1].height = 30

    headers_ws2 = ["Idea", "Modelo", "Veredicto", "Año 1 (USD)", "Año 2 (USD)", "Año 3 (USD)", "Año 4 (USD)", "Año 5 (USD)", "CAGR 5y"]
    ws2.append(headers_ws2)
    ws2.row_dimensions[2].height = 24

    for col_num, h_text in enumerate(headers_ws2, 1):
        cell = ws2.cell(row=2, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    for idx, idea in enumerate(ideas, 1):
        row_num = idx + 2
        rev = idea.get("revenueTimeline", [100000, 300000, 800000, 1500000, 3000000])
        while len(rev) < 5: rev.append(rev[-1] * 1.5)

        ws2.append([
            idea.get("name", ""),
            idea.get("model", ""),
            idea.get("verdict", ""),
            rev[0], rev[1], rev[2], rev[3], rev[4],
            f"=IF(D{row_num}>0, (H{row_num}/D{row_num})^(1/4)-1, 0)"
        ])
        
        ws2.row_dimensions[row_num].height = 20
        fill_row = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
        style_range(ws2, min_col=1, min_row=row_num, max_col=9, max_row=row_num, font=font_body, fill=fill_row)

        for col_c in range(4, 9):
            ws2.cell(row=row_num, column=col_c).number_format = "$#,##0"
        ws2.cell(row=row_num, column=9).number_format = "0.0%"

    # Gráfico Nativo 2: Líneas de Evolución de Ingresos
    chart2 = LineChart()
    chart2.title = "Trayectoria de Ingresos (Año 1 a Año 5)"
    chart2.style = 13
    chart2.y_axis.title = "Ingreso Proyectado (USD)"
    chart2.x_axis.title = "Horizonte Temporal"
    chart2.height = 13
    chart2.width = 20

    data2 = Reference(ws2, min_col=4, min_row=2, max_col=8, max_row=len(ideas)+2)
    cats2 = Reference(ws2, min_col=4, min_row=2, max_col=8) # Encabezados Año 1 a 5 (Fila 2)
    chart2.add_data(data2, titles_from_data=True, from_rows=True)
    chart2.set_categories(cats2)
    ws2.add_chart(chart2, "K2")

    auto_fit_columns(ws2)

    # -------------------------------------------------------------
    # PESTAÑA 3: 03_TAM_SAM_SOM
    # -------------------------------------------------------------
    ws3 = wb.create_sheet(title="03_TAM_SAM_SOM")
    ws3.merge_cells("A1:G1")
    ws3["A1"] = "DIMENSIONAMIENTO DE MERCADO — TAM / SAM / SOM"
    ws3["A1"].font = font_title
    ws3["A1"].alignment = Alignment(vertical="center")

    headers_ws3 = ["Idea", "TAM Global/Reg", "SAM Accesible", "SOM 1 Año", "SOM 3 Años", "SOM 5 Años", "Fuentes Verificadas"]
    ws3.append(headers_ws3)

    for col_num, h_text in enumerate(headers_ws3, 1):
        cell = ws3.cell(row=2, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for idx, idea in enumerate(ideas, 1):
        row_num = idx + 2
        sources = ", ".join([s.get("name", "") for s in idea.get("sourcesList", [])])
        ws3.append([
            idea.get("name", ""),
            idea.get("tam", ""),
            idea.get("sam", ""),
            idea.get("som1y", idea.get("som3y", "")),
            idea.get("som3y", ""),
            idea.get("som5y", idea.get("som3y", "")),
            sources
        ])
        fill_row = fill_zebra if idx % 2 == 0 else PatternFill(fill_type=None)
        style_range(ws3, min_col=1, min_row=row_num, max_col=7, max_row=row_num, font=font_body, fill=fill_row)

    auto_fit_columns(ws3)

    # -------------------------------------------------------------
    # PESTAÑA 4: 04_Unit_Economics
    # -------------------------------------------------------------
    ws4 = wb.create_sheet(title="04_Unit_Economics")
    ws4.merge_cells("A1:H1")
    ws4["A1"] = "UNIT ECONOMICS & CROSS-SELLING POR BUYER PERSONA"
    ws4["A1"].font = font_title
    ws4["A1"].alignment = Alignment(vertical="center")

    headers_ws4 = ["Idea", "Buyer Persona", "CLV Base", "Cross-Sell Boost", "CLV Ajustado", "CAC", "Ratio CLV:CAC", "Payback Period"]
    ws4.append(headers_ws4)

    for col_num, h_text in enumerate(headers_ws4, 1):
        cell = ws4.cell(row=2, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center")

    curr_row = 3
    for idea in ideas:
        name = idea.get("name", "")
        ues = idea.get("unitEconomics", [])
        for ue in ues:
            ws4.append([
                name,
                ue.get("bpName", ""),
                ue.get("clvBase", ""),
                ue.get("crossSellBoost", ""),
                ue.get("clvAdjusted", ""),
                ue.get("cac", ""),
                ue.get("clvCac", ""),
                ue.get("payback", "")
            ])
            style_range(ws4, min_col=1, min_row=curr_row, max_col=8, max_row=curr_row, font=font_body)
            curr_row += 1

    auto_fit_columns(ws4)

    # Guardar archivo Excel
    wb.save(output_path)
    print(f"[OK] Modelo financiero generado en: {output_path}")

def get_default_mock_data():
    return {
        "ideas": [
            {
                "id": "idea-1",
                "name": "SmartB2B Checkout & Credit",
                "model": "B2B SaaS / Transaccional",
                "score": 23,
                "verdict": "PROTOTIPAR",
                "tam": "$2.4B USD",
                "sam": "$340M USD",
                "som1y": "$8.5M USD",
                "som3y": "$42M USD",
                "som5y": "$115M USD",
                "clvCacAdjusted": "6.8:1",
                "buyerPersonas": ["CFO Pyme Industrial", "Director de Ventas B2B"],
                "revenueTimeline": [1200000, 3800000, 8500000, 16000000, 24000000],
                "sourcesList": [{"name": "Banxico"}, {"name": "INEGI"}, {"name": "Statista"}],
                "unitEconomics": [{"bpName": "CFO Pyme", "clvBase": "$4,500 USD", "crossSellBoost": "+$1,800 USD", "clvAdjusted": "$6,300 USD", "cac": "$925 USD", "clvCac": "6.8:1", "payback": "4.5 meses"}]
            },
            {
                "id": "idea-2",
                "name": "Portal Copiloto Fiscal IA",
                "model": "Suscripción B2B SaaS",
                "score": 18,
                "verdict": "VALIDAR MÁS",
                "tam": "$850M USD",
                "sam": "$120M USD",
                "som1y": "$2.8M USD",
                "som3y": "$14M USD",
                "som5y": "$38M USD",
                "clvCacAdjusted": "3.4:1",
                "buyerPersonas": ["Despachos Contables"],
                "revenueTimeline": [800000, 2100000, 4200000, 7800000, 11500000],
                "sourcesList": [{"name": "SAT Informes"}, {"name": "Grand View Research"}],
                "unitEconomics": [{"bpName": "Despacho Contable", "clvBase": "$1,800 USD", "crossSellBoost": "+$400 USD", "clvAdjusted": "$2,200 USD", "cac": "$640 USD", "clvCac": "3.4:1", "payback": "9.6 meses"}]
            },
            {
                "id": "idea-3",
                "name": "Micro-Seguros de Flete",
                "model": "Marketplace",
                "score": 11,
                "verdict": "DESCARTAR",
                "tam": "$420M USD",
                "sam": "$45M USD",
                "som1y": "$600K USD",
                "som3y": "$3.5M USD",
                "som5y": "$9.2M USD",
                "clvCacAdjusted": "1.4:1",
                "buyerPersonas": ["Transportistas Independientes"],
                "revenueTimeline": [180000, 450000, 950000, 1700000, 2600000],
                "sourcesList": [{"name": "AMIS México"}],
                "unitEconomics": [{"bpName": "Transportista", "clvBase": "$350 USD", "crossSellBoost": "+$50 USD", "clvAdjusted": "$400 USD", "cac": "$280 USD", "clvCac": "1.4:1", "payback": "16 meses"}]
            }
        ]
    }

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generador Excel del Dimensionador Estratégico")
    parser.add_argument("--data", help="Ruta a archivo JSON de datos (window.REPORT_DATA)")
    parser.add_argument("--output", default="modelo_financiero_dimensionamiento.xlsx", help="Ruta de salida .xlsx")
    args = parser.parse_args()

    if args.data:
        try:
            with open(args.data, 'r', encoding='utf-8') as f:
                report_data = json.load(f)
        except Exception as e:
            print(f"Error al leer {args.data}: {e}")
            sys.exit(1)
    else:
        report_data = get_default_mock_data()

    generate_excel_model(report_data, args.output)
